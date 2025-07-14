import sys
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import logging
import pyexcel as p
from datetime import datetime
import sys
import os  
import datetime
import pandas as pd

import logging as log
logging.basicConfig(format='%(message)s', level=logging.INFO)

from pathlib import Path

# Absolute path, raw string style:
DATA_DIR = r"C:\Users\davep\Documents\budget\DjangoRestApisPostgreSQL\budget\data"
DATE='Date'
AMOUNT='Amount'
DESCRIPTION = 'Description'

MODEL_DATE='Date'
MODEL_DESCRIPTION= 'Description'
MODEL_AMOUNT= 'Amount'
MODEL_ORIGIN= 'Origin'
MODEL_CATEGORY= 'Category'
def convertFile(fileName="americainExpressStatments/Summary.xls"):
    fileName2 = fileName
    if ".xls"  in fileName2:
        fileName2=fileName2.replace(".xls", ".xlsx")
    p.save_book_as(file_name=fileName,
                dest_file_name=fileName2)
    wb = Workbook()
    ws = wb.active
    wb = load_workbook(fileName2,"rb")
    ws = wb.active
    return ws , fileName2

def findOffsetAndEnd(ws):
    offSetValue="Date"
    offSetRow=-1
    row =1
    maxRow =-1
    trackStart =False         
    while (True):
        char = get_column_letter(1)
        value = ws[char +str(row)].value
        log.debug(f"value  {value} at {char +str(row)}")
       
        if(offSetValue == value):
            offSetRow=row+1
            log.debug(f"offset {offSetRow}")
            trackStart =True
        if(trackStart and None == value):
            maxRow=row
            log.debug(f"offset {offSetRow} maxRow={maxRow} ")
            return offSetRow, maxRow
        
        row +=1
def findDescription(ws,row):
    offSetValue="Description"
    col=1       
    while (True):
        char = get_column_letter(col)
        value = ws[char +str(row-1)].value
        log.debug(f"value  {value} at {char +str(row)}")
       
        if(offSetValue == value):
            offSetRow=row+1
            log.debug(f"Description {col}")
            return col
           
     
        col +=1

 

def findColRows(ws,offSetRow):
    maxCol = 1
    while (True):
        char = get_column_letter(maxCol)
        value = ws[char +str(offSetRow)].value
        log.debug(value)
        if(None == value):

            log.debug(maxCol)
            return maxCol
        maxCol+=1

def printEveryLine(ws):
    offSetRow, maxRow=findOffsetAndEnd(ws)
    for row in  range (offSetRow,maxRow):
        for col in range (1,findColRows(ws,offSetRow)):
            char = get_column_letter(col)
            value= ws[char +str(row)].value
            log.info(f"value  {value} at {char +str(row)}")
        log.info("")

def populateData(ws):
    offSetRow, maxRow=findOffsetAndEnd(ws)
    data=[]
    transactonDescriptCol= findDescription(ws,offSetRow)
    for row in  range (offSetRow,maxRow):
        data.append(createRow(row,ws,transactonDescriptCol))
        log.debug("")

    #convert to dataframe
    df = convertToModels(pd.DataFrame.from_dict(data)) 
    df[DATE] = pd.to_datetime(df[DATE])
    log.debug(df)
    df = df[df[AMOUNT] < 0]
  

    return df

#must be '%Y-%m-%d' to save in datbase
def convertDate(date="11 Jan 2022"):
    return datetime.datetime.strptime(date.replace(".", ""), '%d %b %Y').strftime('%Y-%m-%d')


def createRow(row,ws,transactonDescriptCol) :
    transactonDateCol=1
    amountCol=5


    PAYEMENT = "P"
    SPEND = "S"
    transactonDateCol = get_column_letter(transactonDateCol)
    transactonDescriptLetter = get_column_letter(transactonDescriptCol)
    amountCol = get_column_letter(amountCol)
        

    log.debug("")
 
    transactonDate = ws[ transactonDateCol +str(row)].value
    transactonDescript = ws[transactonDescriptLetter +str(row)].value

    amount= ws[amountCol +str(row)].value
  
    bankAction = PAYEMENT if "-"  in amount else SPEND
    if( "-"  in amount):
        amount = amount.replace("-", "")
    else:
        amount = "-"+amount

    amount= amount.replace("$", "")
    amount = float(amount.replace(',', ''))
  
    
    row = {
    "Date":convertDate(transactonDate),
    "Description":transactonDescript,
    "Amount":amount    }
    log.debug(row)

    return row
# xls is americianEpress
def canParse(full_path):
    return "Summary"  in full_path

def parseByMonth(name=""):
    
    df = parse(name)

    expense_by_month_year = df.groupby([df[DATE].dt.year, df[DATE].dt.month])[AMOUNT].sum()

    # print the aggregated income by month and year
    log.info(expense_by_month_year)

def parseByYear(name=""):
    
    df = parse(name)

   # group the data by month and year, and sum the income
    expense_by_year = df.groupby([df[DATE].dt.year])[AMOUNT].sum()

    # print the aggregated income by month and year
    log.info(expense_by_year)

def convertToModels(df):
 
    new_df = df.loc[:, [DATE,DESCRIPTION, AMOUNT]]

    #basic model 
    #Date  #Description #Amount
    new_df.columns = [MODEL_DATE,MODEL_DESCRIPTION, MODEL_AMOUNT]
    new_df[MODEL_ORIGIN] = 'AmericanExpress'
    
    return new_df

def main(name):

    n = len(sys.argv)
    if(n>1):
        fileName = sys.argv[1]
    else:
        fileName = name
    df =parse(name)
  
def parse(name=""):
    """
    Parses the given file by converting it, populating data from the worksheet,
    and optionally removing the temporary file.
    """
    ws, file_name = convertFile(name)
    data = populateData(ws)

    # Optional: Remove the temporary file if it exists
    try:
        if os.path.exists(file_name):
            os.remove(file_name)
    except Exception as e:
        print(f"Warning: Could not delete temp file '{file_name}': {e}")

    return data


if __name__ == "__main__":
     main(DATA_DIR+"/Summary.xls")