from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import logging
import csv
from datetime import datetime
import pyexcel as p


import logging as log
logging.basicConfig(format='%(message)s', level=logging.DEBUG)


def convertFile(fileName="pcbanking2.csv"):
    
    fileName2 = fileName
    if ".csv"  in fileName2:
        fileName2=fileName2.replace(".csv", ".xlsx")


    wb = Workbook()
    ws = wb.active
    with open(fileName, 'r') as f:
        for row in csv.reader(f):
            ws.append(row)
    wb.save(fileName2)
    wb = load_workbook(fileName2,"rb")
    ws = wb.active
    return ws , fileName2
def convertFile2(fileName="pcbanking2.csv"):
    fileName2 = fileName
    if ".xls"  in fileName2:
        fileName2=fileName2.replace(".xls", ".xlsx")
    p.save_book_as(file_name=fileName,
                dest_file_name=fileName2)
    wb = Workbook()
    ws = wb.active
    wb = load_workbook(fileName2,"rb")
    ws = wb.active
    return ws , fileName2

def findMaxRows(ws):
    maxRow = 1
    while (True):
        char = get_column_letter(1)
        value = ws[char +str(maxRow)].value
        log.debug(value)
        if(None == value):
            maxRow-=1
            log.debug(maxRow)
            return maxRow
        maxRow+=1

def findColRows(ws):
    maxCol = 1
    while (True):
        char = get_column_letter(maxCol)
        value = ws[char +str(1)].value
        log.debug(value)
        if(None == value):
            maxCol-=1
            log.debug(maxCol)
            return maxCol
        maxCol+=1

def printEveryLine(ws):
    for row in  range (1,findMaxRows(ws)):
        for col in range (1,findColRows(ws)):
            char = get_column_letter(col)
            log.info(ws[char +str(row)].value)
        log.info("")

def checkForInvalidCols(ws,fileName):
    row =1
    for col in range (1,findColRows(ws)):
        char = get_column_letter(col)
        gridValue=char +str(row)
        value = ws[gridValue].value
        if(value == '-'):
            log.debug("Invalid col ")
        
            wb = load_workbook(fileName)
            ws = wb.active
            ws.delete_cols(col)
    return (ws)

def printEveryLine(ws):
  
    for row in  range (1,findMaxRows(ws)):
        for col in range (1,findColRows(ws)):
            char = get_column_letter(col)
            log.info(ws[char +str(row)].value)
        log.info("")

def populateData(ws):
    for row in  range (1,findMaxRows(ws)):
        createRow(row,ws)
        log.info("")
def deleteCol(ws,col):
    ws.delete_cols(col)
  

def createRow(row,ws) :
    transactonDateCol = 1
    transactonDescriptCol = 3
    amountCol = 2

    DEPOSIT = "D"
    WITHDRAW = "W"
    transactonDateCol = get_column_letter(transactonDateCol)
    transactonDescriptCol = get_column_letter(transactonDescriptCol)
    amountCol = get_column_letter(amountCol)
        

    log.info("")
 
    transactonDate = ws[ transactonDateCol +str(row)].value
    transactonDescript = ws[transactonDescriptCol +str(row)].value

    amount= ws[amountCol +str(row)].value
  
    bankAction = WITHDRAW if "-"  in amount else DEPOSIT

    amount= float(amount.replace("-", ""))
  
    
    row = {
    "transactonDate":transactonDate,
    "transactonDescript":transactonDescript,
    "amount":amount,
    "bankAction":bankAction,
    }
    log.debug(row)




def main():
    ws, fileName = convertFile()
    ws=checkForInvalidCols(ws,fileName)
    populateData(ws)
if __name__ == "__main__":
    main()
