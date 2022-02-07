from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import logging
import csv
from datetime import datetime
import pyexcel as p


import logging as log
logging.basicConfig(format='%(message)s', level=logging.INFO)

def convertFile(fileName="pcbanking2.csv"):
    fileName2 = fileName
    if ".xls"  in fileName2:
        fileName2=fileName2.replace(".xls", ".xlsx")
    p.save_book_as(file_name=fileName,
                dest_file_name=fileName2)
    wb = Workbook()
    ws = wb.active
    wb = load_workbook(fileName2,"rb")
    ws = wb.active
    return ws , fileName2

def findOffsetAndEnd(ws):
    offSetValue="Date"
    offSetRow=-1
    row =1
    maxRow =-1
    trackStart =False         
    while (True):
        char = get_column_letter(1)
        value = ws[char +str(row)].value
        log.debug(f"value  {value} at {char +str(row)}")
       
        if(offSetValue == value):
            offSetRow=row+1
            log.debug(f"offset {offSetRow}")
            trackStart =True
        if(trackStart and None == value):
            maxRow=row
            log.debug(f"offset {offSetRow} maxRow={maxRow} ")
            return offSetRow, maxRow
        
        row +=1

 

def findColRows(ws,offSetRow):
    maxCol = 1
    while (True):
        char = get_column_letter(maxCol)
        value = ws[char +str(offSetRow)].value
        log.debug(value)
        if(None == value):

            log.debug(maxCol)
            return maxCol
        maxCol+=1

def printEveryLine(ws):
    offSetRow, maxRow=findOffsetAndEnd(ws)
    for row in  range (offSetRow,maxRow):
        for col in range (1,findColRows(ws,offSetRow)):
            char = get_column_letter(col)
            value= ws[char +str(row)].value
            log.info(f"value  {value} at {char +str(row)}")
        log.info("")

def populateData(ws):
    offSetRow, maxRow=findOffsetAndEnd(ws)
    for row in  range (offSetRow,maxRow):
        createRow(row,ws)
        log.info("")




def createRow(row,ws) :
 
    transactonDateCol = 1
    transactonDescriptCol = 3
    amountCol = 4

    PAYEMENT = "P"
    SPEND = "S"
    transactonDateCol = get_column_letter(transactonDateCol)
    transactonDescriptCol = get_column_letter(transactonDescriptCol)
    amountCol = get_column_letter(amountCol)
        

    log.info("")
 
    transactonDate = ws[ transactonDateCol +str(row)].value
    transactonDescript = ws[transactonDescriptCol +str(row)].value

    amount= ws[amountCol +str(row)].value
  
    bankAction = PAYEMENT if "-"  in amount else SPEND
    amount= amount.replace("$", "")
    amount= float(amount.replace("-", ""))
  
    
    row = {
    "transactonDate":transactonDate,
    "transactonDescript":transactonDescript,
    "amount":amount,
    "bankAction":bankAction,
    }
    log.info(row)





def main():
    ws, fileName = convertFile("Summary.xls")
    populateData(ws)
if __name__ == "__main__":
    main()
